import random
from decimal import Decimal
from datetime import datetime, timedelta, date
from typing import List, Dict, Any, Optional, Tuple
import os
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import tkinter as tk
from tkinter import filedialog, messagebox

def get_exports_path() -> str:
    """Get the path to the exports directory."""
    base_dir = os.path.dirname(os.path.abspath(__file__))
    exports_dir = os.path.join(base_dir, 'exports')
    return exports_dir

def ensure_exports_dir() -> None:
    """Ensure the exports directory exists."""
    exports_dir = get_exports_path()
    if not os.path.exists(exports_dir):
        os.makedirs(exports_dir)

def parse_date(date_str: str) -> datetime:
    """Parse a date string into a datetime object."""
    if not isinstance(date_str, str):
        return datetime.now()
        
    formats = [
        "%Y-%m-%d",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d %H:%M:%S"
    ]
    
    for date_format in formats:
        try:
            return datetime.strptime(date_str, date_format)
        except (ValueError, TypeError):
            continue
            
    return datetime.now()  # Fallback to current date if no format matches

def load_routes() -> List[Dict[str, Any]]:
    """Load routes from JSON file."""
    try:
        with open('routes.json', 'r', encoding='utf-8') as f:
            data = json.load(f)
            routes = data.get('routes', [])
            # Convert date strings to datetime objects
            for route in routes:
                if 'date' in route and isinstance(route['date'], str):
                    route['date'] = parse_date(route['date'])
            return routes
    except FileNotFoundError:
        print("Warning: routes.json file not found")
        return []
    except json.JSONDecodeError:
        print("Warning: Invalid JSON format in routes.json")
        return []
    except Exception as e:
        print(f"Error loading routes: {str(e)}")
        return []

def save_routes(routes: List[Dict[str, Any]]) -> None:
    """Save routes to JSON file."""
    # Convert datetime objects to ISO format strings for JSON serialization
    routes_to_save = []
    for route in routes:
        route_copy = route.copy()
        if 'date' in route_copy and isinstance(route_copy['date'], (datetime, date)):
            route_copy['date'] = route_copy['date'].isoformat()
        routes_to_save.append(route_copy)
    
    with open('routes.json', 'w', encoding='utf-8') as f:
        json.dump({'routes': routes_to_save}, f, indent=4, ensure_ascii=False)

def add_route(route: Dict[str, Any]) -> None:
    """Add a new route to JSON file."""
    routes = load_routes()
    # Generate new ID
    max_id = max([r.get('id', 0) for r in routes], default=0)
    route['id'] = max_id + 1
    routes.append(route)
    save_routes(routes)

def update_route(route_id: int, updated_route: Dict[str, Any]) -> bool:
    """Update existing route in JSON file."""
    routes = load_routes()
    for i, route in enumerate(routes):
        if route.get('id') == route_id:
            updated_route['id'] = route_id  # Ensure ID remains the same
            routes[i] = updated_route
            save_routes(routes)
            return True
    return False

def delete_route(route_id: int) -> bool:
    """Delete route from JSON file."""
    routes = load_routes()
    initial_length = len(routes)
    routes = [r for r in routes if r.get('id') != route_id]
    if len(routes) < initial_length:
        save_routes(routes)
        return True
    return False

class Route:
    def __init__(self, name: str = "", start_location: str = "", destination: str = "", 
                 distance: float = 0, trip_purpose: str = ""):
        self.name = name
        self.start_location = start_location
        self.destination = destination
        self.distance = distance
        self.trip_purpose = trip_purpose

    @staticmethod
    def from_dict(route_dict: Dict[str, Any]) -> 'Route':
        """Create a Route instance from a dictionary."""
        try:
            # Parse distance string (e.g., "123.45 km" or "123,45") to float
            distance_str = str(route_dict.get('distance', '0')).replace(',', '.')
            distance = float(''.join(c for c in distance_str if c.isdigit() or c == '.'))
        except (ValueError, TypeError):
            distance = 0.0

        return Route(
            name=str(route_dict.get('name', '')),
            start_location=str(route_dict.get('start_location', '')),
            destination=str(route_dict.get('destination', '')),
            distance=distance,
            trip_purpose=str(route_dict.get('trip_purpose', ''))
        )

def generate_routes_for_segment(target_distance: float, available_routes: List[Dict[str, Any]], 
                              max_routes: int = 5) -> Tuple[List[Route], float]:
    """
    Generate a combination of routes that approximately match the target distance.
    
    Args:
        target_distance: Target distance in kilometers
        available_routes: List of route dictionaries from JSON file
        max_routes: Maximum number of routes to include in the segment
    
    Returns:
        Tuple of (list of selected routes, total distance)
    """
    if not available_routes:
        return [], 0.0

    # Convert all available routes to Route objects
    route_objects = [Route.from_dict(route) for route in available_routes]
    
    # Initialize variables
    selected_routes: List[Route] = []
    total_distance = 0.0
    deviation = 0.05  # Start with 5% deviation
    counter = 0
    max_attempts = 1000
    
    min_target = target_distance * (1 - deviation)
    max_target = target_distance * (1 + deviation)
    
    while total_distance <= min_target and counter < max_attempts:
        counter += 1
        
        # Adjust deviation based on number of attempts
        if counter == 100:
            deviation = 0.10
        elif counter == 200:
            deviation = 0.15
        elif counter == 300:
            deviation = 0.20
        elif counter == 400:
            deviation = 0.25
        elif counter == 500:
            deviation = 0.30
        elif counter == 600:
            deviation = 0.35
        elif counter == 800:
            deviation = 0.50
        elif counter == 900:
            deviation = 1.0
        else:
            min_target = 0
            max_target = 2000
            
        min_target = target_distance * (1 - deviation)
        max_target = target_distance * (1 + deviation)
        
        # Select a random route
        random_route = random.choice(route_objects)
        
        # Check if adding this route would exceed max_routes
        if len(selected_routes) + 1 > max_routes:
            selected_routes = []
            total_distance = 0.0
            continue
            
        # Check if adding this route would exceed max_target
        if total_distance + random_route.distance <= max_target:
            selected_routes.append(random_route)
            total_distance += random_route.distance
            
    # If we couldn't find any routes, return an empty route
    if not selected_routes:
        empty_route = Route("", "", "", 0, "")
        return [empty_route], 0.0
        
    return selected_routes, total_distance

def calculate_route_consumption(distance: float, avg_consumption: float) -> float:
    """Calculate fuel consumption for a given route."""
    return (distance * avg_consumption) / 100

def optimize_routes(routes: List[Dict[str, Any]], initial_tank: float, target_tank: float,
                   avg_consumption: float, refueling_data: Dict[datetime, List[Dict[str, float]]]) -> List[Dict[str, Any]]:
    """
    Optimize routes to match exact fuel balance requirements.
    
    Args:
        routes: Available routes from database
        initial_tank: Initial fuel level
        target_tank: Target final fuel level
        avg_consumption: Average fuel consumption per 100km
        refueling_data: Dictionary of refueling events by date
    
    Returns:
        List of selected routes with dates and fuel states
    """
    # Convert all dates to datetime if they are date objects
    route_dates = []
    for route in routes:
        route_date = route['date']
        if isinstance(route_date, date) and not isinstance(route_date, datetime):
            # Convert date to datetime at midnight
            route_dates.append(datetime.combine(route_date, datetime.min.time()))
        else:
            route_dates.append(route_date)
            
    refuel_dates = list(refueling_data.keys())
    
    # Sort all dates (including both route dates and refueling dates)
    all_dates = sorted(set(route_dates + refuel_dates))
    
    # Initialize tracking variables
    current_tank = initial_tank
    selected_routes = []
    total_consumption = 0
    
    # Create a map of refueling events by date for quick lookup
    refueling_by_date = {}
    for date_val, refuels in refueling_data.items():
        refueling_by_date[date_val] = sum(refuel['amount'] for refuel in refuels)
    
    # Process each day in chronological order
    for current_date in all_dates:
        # First apply any refueling for this day
        if current_date in refueling_by_date:
            current_tank += refueling_by_date[current_date]
        
        # Get available routes for this day
        day_routes = []
        for route in routes:
            route_date = route['date']
            if isinstance(route_date, date) and not isinstance(route_date, datetime):
                route_date = datetime.combine(route_date, datetime.min.time())
            if route_date == current_date:
                day_routes.append(route)
        
        # Try to find a route that can be driven with current fuel level
        for route in day_routes:
            # Calculate fuel needed for this route
            try:
                distance = float(str(route['distance']).replace(',', '.').replace('km', '').strip())
            except (ValueError, TypeError):
                continue
                
            fuel_needed = (distance * avg_consumption) / 100
            
            # Check if we have enough fuel and won't go below target
            if current_tank - fuel_needed >= target_tank:
                selected_routes.append({
                    'date': current_date,
                    'route_name': route.get('name', ''),
                    'start_location': route.get('start_location', ''),
                    'destination': route.get('destination', ''),
                    'distance': distance,
                    'fuel_before': current_tank,
                    'fuel_consumed': fuel_needed,
                    'fuel_after': current_tank - fuel_needed,
                    'purpose': route.get('trip_purpose', 'Služební cesta')
                })
                
                current_tank -= fuel_needed
                total_consumption += fuel_needed
                break
    
    # Verify final fuel state matches target
    final_tank = initial_tank + sum(refueling_by_date.values()) - total_consumption
    if abs(final_tank - target_tank) > 0.001:  # Allow small rounding error
        print(f"Warning: Could not match exact fuel balance. Final: {final_tank:.3f}, Target: {target_tank:.3f}")
        return []
    
    return selected_routes

def integrate_with_ui(initial_tank: float, target_tank: float, start_date: datetime,
                     refueling_data: Dict[datetime, List[Dict[str, float]]],
                     consumption_rate: Decimal) -> Dict[str, Any]:
    """
    Main integration function that handles the optimization process.
    """
    try:
        print("\n--- Generování knihy jízd ---")
        print(f"Počáteční stav nádrže: {initial_tank:.3f} L")
        print(f"Cílový stav nádrže: {target_tank:.3f} L")
        
        # Load routes from JSON
        available_routes = load_routes()
        
        # Add dates to routes based on the month
        month_routes = []
        current_date = datetime.combine(start_date.date(), datetime.min.time())  # Ensure it's a datetime
        
        # Calculate month end date as datetime
        month_start = start_date.replace(day=1)
        next_month = month_start + timedelta(days=32)
        month_end = next_month.replace(day=1) - timedelta(days=1)
        month_end = datetime.combine(month_end.date(), datetime.max.time())  # Set to end of day
        
        # Distribute routes across the month
        route_index = 0
        while current_date <= month_end:
            if route_index < len(available_routes):
                route = available_routes[route_index].copy()
                route['date'] = current_date
                month_routes.append(route)
                route_index += 1
                if route_index >= len(available_routes):
                    route_index = 0
            current_date += timedelta(days=1)

        # Optimize routes with exact fuel balance
        selected_routes = optimize_routes(
            month_routes,
            initial_tank,
            target_tank,
            float(consumption_rate),
            refueling_data
        )
        
        if not selected_routes:
            return {
                'success': False,
                'error': 'Nepodařilo se najít kombinaci tras odpovídající přesné bilanci paliva'
            }

        # Calculate statistics
        total_distance = sum(route['distance'] for route in selected_routes)
        total_consumption = sum(route['fuel_consumed'] for route in selected_routes)
        
        # Calculate total refueled amount
        total_refueled = sum(
            sum(refuel['amount'] for refuel in refuels)
            for refuels in refueling_data.values()
        )

        print(f"\nVýsledky generování:")
        print(f"Počet vygenerovaných tras: {len(selected_routes)}")
        print(f"Celková vzdálenost: {total_distance:.1f} km")
        print(f"Celková spotřeba: {total_consumption:.3f} L")
        print(f"Celkem natankováno: {total_refueled:.3f} L")
        print(f"Konečný stav nádrže: {target_tank:.3f} L")

        return {
            'success': True,
            'solution': selected_routes,
            'statistics': {
                'total_distance': total_distance,
                'total_consumption': total_consumption,
                'total_refueled': total_refueled,
                'final_tank': target_tank
            }
        }
        
    except Exception as e:
        print(f"Error in integrate_with_ui: {str(e)}")
        return {
            'success': False,
            'error': str(e)
        }

def generate_monthly_report(self, month_date, initial_tank, final_tank, avg_consumption, refueling_data, selected_routes):
    """Generate and save a monthly report in Excel format."""
    try:
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Kniha jízd"

        # Define styles
        header_font = Font(bold=True)
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        pink_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Header section
        ws['A1'] = "Kniha jízd"
        ws['B1'] = "Auto:"
        ws['C1'] = self.car_var.get()
        ws['C1'].fill = yellow_fill

        ws['F1'] = "Jméno řidiče:"
        ws['G1'] = "Jiří Kvasnička"
        ws['G1'].fill = yellow_fill

        ws['K1'] = "Principal engineering s.r.o."

        ws['B2'] = "SPZ:"
        ws['C2'] = "EL 148AH"
        ws['C2'].fill = yellow_fill

        ws['F2'] = "Průměrná spotřeba podle TP:"
        ws['G2'] = f"{float(avg_consumption):.1f}"
        ws['G2'].fill = yellow_fill

        # Ensure month_date is datetime
        if isinstance(month_date, str):
            month_date = parse_date(month_date)
        elif isinstance(month_date, date) and not isinstance(month_date, datetime):
            month_date = datetime.combine(month_date, datetime.min.time())

        # Date range
        ws['A3'] = "Sledované období od"
        ws['D3'] = month_date.strftime("%d.%m.%Y")
        ws['A4'] = "Sledované období do"
        month_end = (month_date.replace(day=1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        if isinstance(month_end, date) and not isinstance(month_end, datetime):
            month_end = datetime.combine(month_end, datetime.max.time())
        ws['D4'] = month_end.strftime("%d.%m.%Y")

        # Calculate total distance from selected routes
        total_distance = sum(route['distance'] for route in selected_routes)
        
        # Tachometer states - calculate based on total distance
        initial_tachometer = 94439  # This should come from database or user input
        final_tachometer = initial_tachometer + int(total_distance)
        
        ws['A5'] = "Počáteční stav tachometru:"
        ws['B5'] = str(initial_tachometer)
        ws['B5'].fill = yellow_fill

        ws['A6'] = "Konečný stav tachometru:"
        ws['B6'] = str(final_tachometer)
        ws['B6'].fill = yellow_fill

        # Fuel states
        ws['A7'] = "Stav nádrže k poslednímu dni předch.měsíce (l,Kč/l)"
        ws['B7'] = f"{initial_tank:.0f}"
        ws['B7'].fill = yellow_fill

        ws['A8'] = "Stav nádrže k poslednímu dni v aktuálním měsíci (l, Kč/l)"
        ws['B8'] = f"{final_tank:.0f}"
        ws['B8'].fill = yellow_fill

        # Right side information
        ws['E3'] = "Počet km - tachometr"
        ws['F3'] = str(int(total_distance))
        ws['F3'].fill = yellow_fill

        ws['E4'] = "Počet km - kniha jízd"
        ws['F4'] = str(int(total_distance))
        ws['F4'].fill = yellow_fill

        ws['E5'] = "Spotřeba PHM pro soukromé jízdy v tuz. (Kč)"
        ws['F5'] = "0,00"

        ws['E6'] = "Spotřeba PHM pro soukromé jízdy v zahr.(Kč)"
        ws['F6'] = "0,00"

        # Calculate average consumption for the period
        total_fuel_used = sum(route['fuel_consumed'] for route in selected_routes)
        avg_period_consumption = (total_fuel_used * 100) / total_distance if total_distance > 0 else 0

        ws['E7'] = "Průměrná spotřeba v období: ( l/100 km)"
        ws['F7'] = f"{avg_period_consumption:.2f}"
        ws['F7'].fill = pink_fill

        ws['E8'] = "Průměrná cena PHM tuz.(Kč bez DPH)"
        ws['F8'] = "0,00"

        ws['E9'] = "Zahraniční cena"
        ws['F9'] = "0,00"

        # Routes table headers
        current_row = 11
        headers = ["Trasa", "Účel", "Datum", "Km", "služebně", "soukr.tuz", "soukr.zahr", 
                  "počet litrů/kW (ČR)", "Kč bez DPH (ČR)", "počet litrů (zahr)", "cena v Kč (zahr.)"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.font = header_font
            cell.border = border

        # Add routes data
        current_row += 1
        
        # Ensure all dates are datetime before sorting
        routes_to_sort = []
        for route in selected_routes:
            route_copy = route.copy()
            if isinstance(route_copy['date'], str):
                route_copy['date'] = parse_date(route_copy['date'])
            elif isinstance(route_copy['date'], date) and not isinstance(route_copy['date'], datetime):
                route_copy['date'] = datetime.combine(route_copy['date'], datetime.min.time())
            routes_to_sort.append(route_copy)
            
        # Sort routes by date
        sorted_routes = sorted(routes_to_sort, key=lambda x: x['date'])
        
        for route in sorted_routes:
            ws.cell(row=current_row, column=1, value=f"{route['start_location']} - {route['destination']}")
            ws.cell(row=current_row, column=2, value=route['purpose'])
            ws.cell(row=current_row, column=3, value=route['date'].strftime("%A %d. %B %Y"))
            ws.cell(row=current_row, column=4, value=route['distance'])
            ws.cell(row=current_row, column=5, value=route['distance'])  # All km are business
            ws.cell(row=current_row, column=8, value=route['fuel_consumed'])
            current_row += 1

        # Add totals row
        ws.cell(row=current_row, column=1, value="Celkem")
        ws.cell(row=current_row, column=4, value=f"=SUM(D12:D{current_row-1})")
        ws.cell(row=current_row, column=5, value=f"=SUM(E12:E{current_row-1})")
        ws.cell(row=current_row, column=8, value=f"=SUM(H12:H{current_row-1})")

        # Adjust column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 15

        # Create exports directory if it doesn't exist
        ensure_exports_dir()
        
        # Save the file
        month_year = month_date.strftime("%Y-%m")
        filename = f"kniha_jizd_{month_year}.xlsx"
        filepath = os.path.join(get_exports_path(), filename)
        
        wb.save(filepath)
        
        # Show success message and ask if user wants to open the file
        if messagebox.askyesno(
            "Report vytvořen",
            f"Report byl úspěšně vytvořen jako {filename}.\nChcete ho otevřít?",
            parent=self.window
        ):
            try:
                os.startfile(filepath)
            except Exception as e:
                messagebox.showwarning(
                    "Upozornění",
                    f"Report byl vytvořen, ale nepodařilo se ho automaticky otevřít.\nNajdete ho ve složce exports.\nDetail: {str(e)}",
                    parent=self.window
                )

    except Exception as e:
        messagebox.showerror(
            "Chyba",
            f"Nepodařilo se vygenerovat report: {str(e)}",
            parent=self.window
        )
        print(f"Error generating report: {str(e)}")
