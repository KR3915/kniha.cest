import tkinter as tk
import customtkinter as ctk
from tkinter import ttk, simpledialog, messagebox
import requests
import database # Předpokládá se, že máte soubor database.py
import json
import math
import random
import time # Pro simulaci časového zpoždění, pokud je potřeba
import logo_utils
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Side, Border
from datetime import datetime, timedelta, date
from tkcalendar import DateEntry  # We'll need to install this package
import os
import subprocess
from export_utils import get_exports_path, open_exports_folder, ensure_exports_dir
import heapq  # Přidáváme chybějící import
from fuel_optimizer import integrate_with_ui
from decimal import Decimal

# --- Styling Constants ---
PADDING = 20
CORNER_RADIUS = 8

# Define font settings
FONT_FAMILY = "Roboto"
FONT_LARGE = ("Roboto", 24, "bold")
FONT_MEDIUM = ("Roboto", 14, "bold")
FONT_SMALL = ("Roboto", 12, "bold")
FONT_NORMAL = ("Roboto", 12)

# Configure default font for all widgets
ctk.ThemeManager.theme["CTkFont"]["family"] = FONT_FAMILY

# TomTom API Key
TOMTOM_API_KEY = "Guh742xz9ZSxx11iki85pe5bvprH9xL9"

# Global variables
user_root = None
current_username = "Unknown User"
current_user_id = None
user_routes_tree = None
all_user_routes_cached = []
fuel_filter = "Vše"  # Default value for fuel filter
current_sort_column = "name"  # Default sort column
current_sort_order = "asc"  # Default sort order

# Global variables for entries
name_entry = None
start_entry = None
dest_entry = None
purpose_entry = None

# Global variables for gas station search
gas_station_entry = None
gas_station_result_label = None

# Global variable for progress label
progress_label = None

# Callback to return to login screen
_create_login_widgets_callback = None

# Set the appearance mode and default color theme
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

# --- Helper Functions ---

def get_coordinates(location_name):
    """
    Uses TomTom Geocoding API to get coordinates for a given location name.
    """
    if not TOMTOM_API_KEY or TOMTOM_API_KEY == "YOUR_TOMTOM_API_KEY_HERE": # Added a check for default placeholder
        messagebox.showwarning("API Klíč chybí", "Prosím, zadejte svůj platný TomTom API klíč do souboru user_screen.py", parent=user_root)
        return None

    geocode_url = f"https://api.tomtom.com/search/2/geocode/{location_name}.json?key={TOMTOM_API_KEY}"
    try:
        response = requests.get(geocode_url)
        response.raise_for_status() # Raise an exception for HTTP errors (4xx or 5xx)
        data = response.json()

        print(f"DEBUG: Geocode API Response for '{location_name}': {json.dumps(data, indent=2)}") # Debug print

        if data and 'results' in data and len(data['results']) > 0:
            position = data['results'][0]['position']
            return f"{position['lat']},{position['lon']}" # Returns "latitude,longitude" string
        else:
            # If no results, check if TomTom API provided an error message
            error_message = data.get('error', data.get('detailedError', 'Neznámá chyba'))
            messagebox.showerror(
                "Chyba geokódování",
                f"Nepodařilo se najít souřadnice pro '{location_name}'. Zkuste přesnější název.\n"
                f"API zpráva: {error_message}",
                parent=user_root
            )
            return None
    except requests.exceptions.HTTPError as e:
        # Specific handling for HTTP errors (e.g., 401 Unauthorized, 403 Forbidden, 429 Too Many Requests)
        messagebox.showerror(
            "Chyba sítě (HTTP)",
            f"Chyba při komunikaci s TomTom Geocoding API pro '{location_name}':\n"
            f"Status Kód: {e.response.status_code}\n"
            f"Zpráva: {e.response.text}\n"
            f"Zkontrolujte API klíč a internetové připojení.",
            parent=user_root
        )
        return None
    except requests.exceptions.ConnectionError as e:
        messagebox.showerror("Chyba sítě (připojení)", f"Nelze se připojit k TomTom API. Zkontrolujte internetové připojení.\nChyba: {e}", parent=user_root)
        return None
    except requests.exceptions.Timeout:
        messagebox.showerror("Chyba sítě (timeout)", "Vypršel časový limit pro připojení k TomTom API. Zkuste to znovu.", parent=user_root)
        return None
    except requests.exceptions.RequestException as e:
        messagebox.showerror("Chyba sítě", f"Obecná chyba při komunikaci s TomTom Geocoding API: {e}", parent=user_root)
        return None
    except json.JSONDecodeError:
        messagebox.showerror("Chyba dat", f"Nelze dekódovat odpověď z TomTom API pro '{location_name}'. Možná poškozená data.", parent=user_root)
        return None


def get_reverse_geocode_location(latitude, longitude):
    """
    Uses TomTom Reverse Geocoding API to get a location name from coordinates.
    """
    if not TOMTOM_API_KEY:
        return "Neznámá lokace (chybí API klíč)"

    reverse_geocode_url = f"https://api.tomtom.com/search/2/reverseGeocode/{latitude},{longitude}.json?key={TOMTOM_API_KEY}"
    try:
        response = requests.get(reverse_geocode_url)
        response.raise_for_status()
        data = response.json()
        if data and 'addresses' in data and len(data['addresses']) > 0:
            address = data['addresses'][0]['address']
            # Prioritize street and number, then municipality, then country
            if 'streetName' in address and 'streetNumber' in address:
                return f"{address['streetName']} {address['streetNumber']}, {address.get('municipality', '')}"
            elif 'municipality' in address:
                return f"{address['municipality']}, {address.get('country', '')}"
            else:
                return address.get('freeformAddress', f"{latitude},{longitude}")
        else:
            return f"Neznámá lokace ({latitude},{longitude})"
    except requests.exceptions.RequestException:
        return f"Neznámá lokace (chyba API: {latitude},{longitude})"


def get_route_data(start_coords, destination_coords):
    """
    Uses TomTom Routing API to get route details between two coordinate sets.
    Returns distance in meters, travel time in seconds, and legs for waypoint extraction.
    """
    if not TOMTOM_API_KEY or TOMTOM_API_KEY == "YOUR_TOMTOM_API_KEY_HERE":
        messagebox.showwarning("API Klíč chybí", "Prosím, zadejte svůj platný TomTom API klíč do souboru user_screen.py", parent=user_root)
        return None

    route_url = (
        f"https://api.tomtom.com/routing/1/calculateRoute/{start_coords}:{destination_coords}/json?"
        f"key={TOMTOM_API_KEY}&travelMode=car&instructionsType=text&language=cs-CZ&routeType=fastest&traffic=false&avoid=unpavedRoads"
    )
    try:
        response = requests.get(route_url)
        response.raise_for_status()
        data = response.json()

        print(f"DEBUG: Route API Response from {start_coords} to {destination_coords}: {json.dumps(data, indent=2)}") # Debug print

        if data and 'routes' in data and len(data['routes']) > 0:
            route = data['routes'][0]['summary']
            legs = data['routes'][0]['legs'] # Get legs for waypoints
            
            print(f"DEBUG: Route found. Distance: {route['lengthInMeters']}m, Time: {route['travelTimeInSeconds']}s")
            return {
                'distance': route['lengthInMeters'],
                'travel_time': route['travelTimeInSeconds'],
                'legs': legs # Pass legs for waypoint extraction
            }
        else:
            # If no routes, check if TomTom API provided an error message
            error_message = data.get('error', data.get('detailedError', 'Neznámá chyba'))
            messagebox.showerror(
                "Chyba routování",
                f"Nepodařilo se najít trasu mezi {start_coords} a {destination_coords}.\n"
                f"API zpráva: {error_message}",
                parent=user_root
            )
            print(f"DEBUG: No route found from {start_coords} to {destination_coords}. API Response: {json.dumps(data, indent=2)}")
            return None
    except requests.exceptions.HTTPError as e:
        # Specific handling for HTTP errors
        messagebox.showerror(
            "Chyba sítě (HTTP)",
            f"Chyba při komunikaci s TomTom Routing API:\n"
            f"Status Kód: {e.response.status_code}\n"
            f"Zpráva: {e.response.text}\n"
            f"Zkontrolujte API klíč a internetové připojení.",
            parent=user_root
        )
        return None
    except requests.exceptions.ConnectionError as e:
        messagebox.showerror("Chyba sítě (připojení)", f"Nelze se připojit k TomTom API. Zkontrolujte internetové připojení.\nChyba: {e}", parent=user_root)
        return None
    except requests.exceptions.Timeout:
        messagebox.showerror("Chyba sítě (timeout)", "Vypršel časový limit pro připojení k TomTom API. Zkuste to znovu.", parent=user_root)
        return None
    except requests.exceptions.RequestException as e:
        messagebox.showerror("Chyba sítě", f"Obecná chyba při komunikaci s TomTom Routing API: {e}", parent=user_root)
        return None
    except json.JSONDecodeError:
        messagebox.showerror("Chyba dat", f"Nelze dekódovat odpověď z TomTom API pro routování. Možná poškozená data.", parent=user_root)
        return None

def calculate_fuel_consumption(distance_km, avg_consumption_per_100km=7.0):
    """
    Calculates estimated fuel consumption for a given distance.
    :param distance_km: Distance in kilometers.
    :param avg_consumption_per_100km: Average fuel consumption in liters per 100km.
    :return: Estimated fuel consumption in liters.
    """
    return (distance_km / 100) * avg_consumption_per_100km

def get_gas_stations_along_route(waypoints, radius_km=5, limit=5):
    """
    Simulates finding gas stations near the route using waypoints.
    In a real app, this would use a POI search API with a route geometry.
    For now, it randomly picks from a predefined list.
    """
    # This is a highly simplified simulation.
    # In a real application, you would use TomTom Search API's along route search:
    # https://developer.tomtom.com/search-api/documentation/search-api/poi-search-along-route
    
    # Example simulated gas stations (replace with actual API calls)
    simulated_gas_stations = [
        {"name": "OMV Praha", "coords": [50.081, 14.425]},
        {"name": "Shell Brno", "coords": [49.195, 16.608]},
        {"name": "Benzina Plzen", "coords": [49.738, 13.376]},
        {"name": "Mol Ostrava", "coords": [49.820, 18.262]},
        {"name": "EuroOil Olomouc", "coords": [49.594, 17.251]},
    ]
    
    found_stations = []
    # In a real scenario, you'd iterate through waypoints or use TomTom's along route search.
    # For this simulation, we'll just return a few random ones if waypoints exist.
    if waypoints:
        random.shuffle(simulated_gas_stations)
        found_stations = simulated_gas_stations[:min(len(simulated_gas_stations), limit)]
    
    return found_stations

def save_route_to_db(route_name, start_loc, dest_loc, distance, travel_time, fuel_consumption, gas_stations, needs_fuel, waypoints, trip_purpose):
    """Saves a new route to the database."""
    if database.add_route(
        current_user_id,  # user_id
        route_name,      # name
        start_loc,       # start_location
        dest_loc,        # destination
        f"{distance/1000:.2f} km",  # distance in km, formatted
        travel_time,     # travel_time
        fuel_consumption,  # fuel_consumption
        gas_stations,    # gas_stations
        needs_fuel,      # needs_fuel
        trip_purpose,    # trip_purpose
        date.today(),    # route_date
        waypoints       # waypoints
    ):
        messagebox.showinfo("Úspěch", "Trasa byla úspěšně uložena!", parent=user_root)
        load_user_routes() # Refresh the list
    else:
        messagebox.showerror("Chyba", "Nepodařilo se uložit trasu.", parent=user_root)

def update_route_in_db(route_id, user_id, name, start_location, destination, distance, travel_time, fuel_consumption, gas_stations, needs_fuel, waypoints, trip_purpose):
    """Updates an existing route in the database."""
    if database.update_route(
        route_id, user_id, name, start_location, destination,
        distance, travel_time, fuel_consumption,
        gas_stations, needs_fuel, waypoints, trip_purpose
    ):
        messagebox.showinfo("Úspěch", "Trasa byla úspěšně aktualizována!", parent=user_root)
        load_user_routes() # Refresh the list
    else:
        messagebox.showerror("Chyba", "Nepodařilo se aktualizovat trasu.", parent=user_root)

def delete_route_from_db(route_id):
    """Deletes a route from the database."""
    if messagebox.askyesno("Smazat trasu", "Opravdu chcete smazat tuto trasu?", parent=user_root):
        if database.delete_route(route_id):
            messagebox.showinfo("Úspěch", "Trasa byla úspěšně smazána.", parent=user_root)
            load_user_routes() # Refresh the list
        else:
            messagebox.showerror("Chyba", "Nepodařilo se smazat trasu.", parent=user_root)


# --- UI Callbacks ---

def calculate_route(route_name, start_location, destination, trip_purpose, progress_label=None):
    """Calculate route between two points and save it to database."""
    try:
        if not route_name:
            messagebox.showwarning("Upozornění", "Zadejte název trasy.", parent=user_root)
        return
        if not start_location:
            messagebox.showwarning("Upozornění", "Zadejte výchozí místo.", parent=user_root)
        return
        if not destination:
            messagebox.showwarning("Upozornění", "Zadejte cílové místo.", parent=user_root)
            return
        if progress_label:
            progress_label.configure(text="Získávám souřadnice, prosím čekejte...")
            user_root.update_idletasks()

        start_coords_str = get_coordinates(start_location)
        if not start_coords_str:
            if progress_label:
                progress_label.configure(text="")
            return
        destination_coords_str = get_coordinates(destination)
        if not destination_coords_str:
            if progress_label:
                progress_label.configure(text="")
            return
        if progress_label:
            progress_label.configure(text="Počítám trasu, prosím čekejte...")
            user_root.update_idletasks()
        route_info = get_route_data(start_coords_str, destination_coords_str)
        if not route_info:
            if progress_label:
                progress_label.configure(text="")
            messagebox.showerror("Chyba", "Nepodařilo se vypočítat trasu.", parent=user_root)
            return
        distance = route_info['distance']
        travel_time = route_info['travel_time']
        waypoints_list = []
        if 'legs' in route_info and route_info['legs']:
            for leg in route_info['legs']:
                if 'points' in leg:
                    for point in leg['points']:
                        waypoints_list.append([point['latitude'], point['longitude']])
        fuel_consumption = calculate_fuel_consumption(distance / 1000)
        gas_stations = get_gas_stations_along_route(waypoints_list)
        needs_fuel = bool(gas_stations)
        save_route_to_db(
            route_name, start_location, destination,
            distance, travel_time, fuel_consumption,
            gas_stations, needs_fuel, waypoints_list,
            trip_purpose
        )
        name_entry.delete(0, 'end')
        start_entry.delete(0, 'end')
        dest_entry.delete(0, 'end')
        purpose_entry.delete(0, 'end')
    except Exception as e:
        print(f"Error in calculate_route: {str(e)}")
        messagebox.showerror("Chyba", f"Nastala chyba při výpočtu trasy: {str(e)}", parent=user_root)
        if progress_label:
            progress_label.configure(text="")


def load_user_routes():
    """Loads routes for the current user from the database and populates the Treeview."""
    global all_user_routes_cached
    all_user_routes_cached = []  # Clear cache

    # Clear existing items
    for item in user_routes_tree.get_children():
        user_routes_tree.delete(item)

    routes = database.get_routes_by_user(current_user_id)
    if routes:
        for route_data in routes:
            # JSONB columns are already deserialized by psycopg2
            # Just ensure we have lists even if the fields are None
            route_data['gas_stations'] = route_data['gas_stations'] if route_data['gas_stations'] is not None else []
            route_data['waypoints'] = route_data['waypoints'] if route_data['waypoints'] is not None else []

            # Add to cache
            all_user_routes_cached.append(route_data)
    
            # Add to Treeview
            needs_fuel_display = "Ano" if route_data.get('needs_fuel', False) else "Ne"
            num_gas_stations = len(route_data.get('gas_stations', []))

            display_values = (
                str(route_data.get('name', '')),
                str(route_data.get('trip_purpose', '')),  # Use trip_purpose here
                str(route_data.get('start_location', '')),
                str(route_data.get('destination', '')),
                str(route_data.get('distance', '')),
                f"{route_data.get('travel_time', 0)//3600}h {(route_data.get('travel_time', 0)%3600)//60}m",
                f"{route_data.get('fuel_consumption', 0.0):.1f}",
                str(num_gas_stations),
                needs_fuel_display
            )
            user_routes_tree.insert("", "end", values=display_values, tags=(route_data['id'],))


def edit_route_button_click():
    """Handles editing a selected route."""
    selected_item = user_routes_tree.focus()
    if not selected_item:
        messagebox.showwarning("Úprava trasy", "Vyberte trasu, kterou chcete upravit.", parent=user_root)
        return

    # Get route ID from the selected item
    route_id = user_routes_tree.item(selected_item, "tags")[0]  # Assuming ID is stored as a tag

    # Find the full route data from cache
    selected_route_data = next((r for r in all_user_routes_cached if str(r['id']) == str(route_id)), None)

    if not selected_route_data:
        messagebox.showerror("Chyba", "Nepodařilo se načíst data pro úpravu trasy.", parent=user_root)
        return

    # Show edit dialog
    dialog = EditRouteDialog(user_root, selected_route_data)
    user_root.wait_window(dialog.window)
    
    # If user cancelled, return
    if not dialog.result:
        return
        
    # Get the edited values
    new_name = dialog.result['name']
    new_start_location = dialog.result['start_location']
    new_destination = dialog.result['destination']
    trip_purpose = dialog.result['trip_purpose']

    # If start or destination changed, recalculate the route
    recalculate = False
    if new_start_location != selected_route_data['start_location'] or new_destination != selected_route_data['destination']:
        recalculate = True
        
        if progress_label:
            progress_label.configure(text="Přepočítávám trasu, prosím čekejte...")
            user_root.update_idletasks()

        start_coords_str = get_coordinates(new_start_location)
        if not start_coords_str:
            if progress_label: progress_label.configure(text="")
            return
        destination_coords_str = get_coordinates(new_destination)
        if not destination_coords_str:
            if progress_label: progress_label.configure(text="")
            return

        recalculated_route_info = get_route_data(start_coords_str, destination_coords_str)
        if progress_label: progress_label.configure(text="")

        if not recalculated_route_info:
            messagebox.showerror("Chyba", "Nepodařilo se přepočítat trasu s novými údaji.", parent=user_root)
            return
        
        updated_distance = recalculated_route_info['distance']
        updated_travel_time = recalculated_route_info['travel_time']
        
        updated_waypoints_list = []
        if 'legs' in recalculated_route_info and recalculated_route_info['legs']:
            for leg in recalculated_route_info['legs']:
                if 'points' in leg:
                    for point in leg['points']:
                        updated_waypoints_list.append([point['latitude'], point['longitude']])

        updated_fuel_consumption = calculate_fuel_consumption(updated_distance / 1000)
        updated_gas_stations = get_gas_stations_along_route(updated_waypoints_list)
        updated_needs_fuel = bool(updated_gas_stations)

    else:  # No need to recalculate, use existing data
        try:
            updated_distance = float(str(selected_route_data['distance']).replace(' km', '').replace(',', '.')) * 1000
        except ValueError:
            updated_distance = 0.0  # Default if conversion fails
        
        updated_travel_time = selected_route_data['travel_time']
        updated_fuel_consumption = selected_route_data['fuel_consumption']
        updated_gas_stations = selected_route_data['gas_stations']
        updated_needs_fuel = selected_route_data['needs_fuel']
        updated_waypoints_list = selected_route_data['waypoints']

    update_route_in_db(
        route_id, current_user_id, new_name, new_start_location, new_destination,
        updated_distance, updated_travel_time, updated_fuel_consumption,
        updated_gas_stations, updated_needs_fuel, updated_waypoints_list,
        trip_purpose
    )


def delete_route_button_click():
    """Handles deleting a selected route."""
    selected_item = user_routes_tree.focus()
    if not selected_item:
        messagebox.showwarning("Smazat trasu", "Vyberte trasu, kterou chcete smazat.", parent=user_root)
        return
    
    route_id = user_routes_tree.item(selected_item, "tags")[0] # Assuming ID is stored as a tag
    delete_route_from_db(route_id)

# --- Filtering and Sorting Logic ---

def apply_filters_and_sort():
    """Applies current filters and sorting to the cached route data and updates the Treeview."""
    # Clear existing entries in the Treeview
    for item in user_routes_tree.get_children():
        user_routes_tree.delete(item)

    filtered_routes = []

    for route in all_user_routes_cached:
        match_name = name_entry.get().strip().lower() in route.get('name', '').lower()
        match_start = start_entry.get().strip().lower() in route.get('start_location', '').lower()
        match_destination = dest_entry.get().strip().lower() in route.get('destination', '').lower()

        match_fuel = True
        if fuel_filter == "Ano":
            match_fuel = route.get('needs_fuel', False)
        elif fuel_filter == "Ne":
            match_fuel = not route.get('needs_fuel', False)
        # If fuel_filter is "Vše", match_fuel remains True
        
        if match_name and match_start and match_destination and match_fuel:
            filtered_routes.append(route)

    # Sort the filtered routes
    def get_sort_key(route):
        val = route.get(current_sort_column)
        if current_sort_column == 'distance':
            # Handle cases where distance might be None or not a string
            if val is None: return -1 # Or some other default for sorting
            try:
                return float(str(val).replace(' km', '').replace(',', '.')) # Convert "123.45 km" to 123.45 for sorting
            except ValueError:
                return -1 # Handle cases where conversion fails
        elif current_sort_column == 'travel_time':
            return val # Already integer
        elif current_sort_column == 'fuel_consumption':
            return val # Already float
        elif current_sort_column == 'needs_fuel':
            return 1 if val else 0 # Sort boolean
        return str(val).lower() # Default to string comparison for others

    filtered_routes.sort(key=get_sort_key, reverse=(current_sort_order == "desc"))

    # Populate Treeview with sorted and filtered data
    for route in filtered_routes:
        needs_fuel_display = "Ano" if route.get('needs_fuel', False) else "Ne"
        num_gas_stations = len(route.get('gas_stations', []))

        # Ensure all values are strings before inserting into Treeview
        display_values = (
            str(route.get('name', '')),
            str(route.get('trip_purpose', '')),  # Use trip_purpose here
            str(route.get('start_location', '')),
            str(route.get('destination', '')),
            str(route.get('distance', '')),
            f"{route.get('travel_time', 0)//3600}h {(route.get('travel_time', 0)%3600)//60}m",
            f"{route.get('fuel_consumption', 0.0):.1f}",
            str(num_gas_stations),
            needs_fuel_display
        )
        user_routes_tree.insert("", "end", values=display_values, tags=(route['id'],)) # Store route ID as a tag for easy retrieval


def set_sort_and_apply(column, order):
    """Sets the global sort column and order, then applies filters and sorting."""
    global current_sort_column, current_sort_order
    current_sort_column = column
    current_sort_order = order
    apply_filters_and_sort()

def go_back_to_login():
    """Destroys current screen content and returns to login using the stored callback."""
    global user_root, _create_login_widgets_callback, all_user_routes_cached

    # Clear current screen widgets
    for widget in user_root.winfo_children():
        widget.destroy()

    # Reset window size and title
    user_root.geometry("300x200")
    user_root.title("Přihlašovací obrazovka")

    # Clear cached routes
    all_user_routes_cached = []

    # Call the callback to recreate login widgets
    if _create_login_widgets_callback:
        _create_login_widgets_callback()


# --- Main UI Setup ---

def show_user_page(app_main_window, username, user_id, create_login_widgets_callback_func):
    """Creates and displays the user interface within the provided app_main_window."""
    global user_root, current_username, current_user_id, _create_login_widgets_callback
    global user_routes_tree, progress_label
    global name_entry, start_entry, dest_entry, purpose_entry
    global gas_station_entry, gas_station_result_label

    # Use the passed main window directly instead of creating a Toplevel
    user_root = app_main_window 
    # Clear any existing widgets from the main window (e.g., login widgets)
    # This should have been done in login_screen.py before calling this function,
    # but as a safeguard, we can do it here too or ensure it's consistently handled.
    # For now, assuming login_screen.py clears it before calling show_user_page.

    user_root.title(f"Kniha.cest - {username}")
    user_root.geometry("1200x800") # Set geometry for the user page
    # Center the window if it's being reshaped
    user_root.update_idletasks()
    x = user_root.winfo_screenwidth() // 2 - user_root.winfo_width() // 2
    y = user_root.winfo_screenheight() // 2 - user_root.winfo_height() // 2
    user_root.geometry(f"+{x}+{y}")
    user_root.resizable(True, True) # User page can be resizable

    # Ensure the logout callback is set for the window's close button
    user_root.protocol("WM_DELETE_WINDOW", go_back_to_login)

    current_username = username
    current_user_id = user_id
    _create_login_widgets_callback = create_login_widgets_callback_func

    # Main container - now directly in user_root (which is app_main_window)
    main_container = ctk.CTkFrame(user_root)
    main_container.pack(fill="both", expand=True, padx=PADDING, pady=PADDING)

    # Header frame
    header_frame = ctk.CTkFrame(main_container, fg_color="transparent")
    header_frame.pack(fill="x", pady=(0, PADDING))

    # Add logo at the top
    logo_label = logo_utils.setup_logo(header_frame)
    if logo_label:
        logo_label.pack(side="left", padx=(0, PADDING))

    # Theme switch in header
    theme_switch = ctk.CTkSwitch(
        header_frame,
        text="Tmavý režim",
        command=toggle_theme,
        onvalue="dark",
        offvalue="light",
        font=FONT_SMALL
    )
    theme_switch.pack(side="right", padx=PADDING)

    # Logout button in header
    logout_button = ctk.CTkButton(
        header_frame,
        text="Odhlásit se",
        command=go_back_to_login,
        width=120,
        fg_color="transparent",
        border_width=2,
        text_color=("gray10", "gray90"),
        font=FONT_SMALL
    )
    logout_button.pack(side="right", padx=PADDING)

    # Content frame with two columns
    content_frame = ctk.CTkFrame(main_container)
    content_frame.pack(fill="both", expand=True)

    # Left column - Route creation
    left_column = ctk.CTkFrame(content_frame, fg_color="transparent")
    left_column.pack(side="left", fill="both", expand=True, padx=(0, PADDING))

    route_creation_frame = ctk.CTkFrame(left_column)
    route_creation_frame.pack(fill="x", pady=(0, PADDING))

    # Route creation title
    route_title = ctk.CTkLabel(
        route_creation_frame,
        text="Vytvořit novou trasu",
        font=FONT_MEDIUM
    )
    route_title.pack(pady=(PADDING, 0))

    # Route name entry
    name_label = ctk.CTkLabel(route_creation_frame, text="Název trasy:", font=FONT_NORMAL)
    name_label.pack(pady=(0, 5), anchor="w", padx=20)
    
    name_entry = ctk.CTkEntry(route_creation_frame, width=300)
    name_entry.pack(pady=(0, 10), padx=20)

    # Add trip purpose field
    purpose_label = ctk.CTkLabel(route_creation_frame, text="Účel cesty:", font=FONT_NORMAL)
    purpose_label.pack(pady=(0, 5), anchor="w", padx=20)
    purpose_entry = ctk.CTkEntry(route_creation_frame, width=300)
    purpose_entry.pack(pady=(0, 10), padx=20)

    # Start location entry
    start_label = ctk.CTkLabel(route_creation_frame, text="Počáteční bod:", font=FONT_SMALL)
    start_label.pack(anchor="w", padx=PADDING)
    
    start_entry = ctk.CTkEntry(
        route_creation_frame,
        placeholder_text="Zadejte počáteční bod",
        width=300,
        font=FONT_SMALL
    )
    start_entry.pack(padx=PADDING, pady=(5, PADDING))

    # Destination entry
    dest_label = ctk.CTkLabel(route_creation_frame, text="Cílový bod:", font=FONT_SMALL)
    dest_label.pack(anchor="w", padx=PADDING)
    
    dest_entry = ctk.CTkEntry(
        route_creation_frame,
        placeholder_text="Zadejte cílový bod",
        width=300,
        font=FONT_SMALL
    )
    dest_entry.pack(padx=PADDING, pady=(5, PADDING))

    # Calculate route button
    calculate_button = ctk.CTkButton(
        route_creation_frame,
        text="Vypočítat trasu",
        command=calculate_route,
        width=300,
        font=FONT_SMALL
    )
    calculate_button.pack(pady=PADDING)

    # Progress label
    progress_label = ctk.CTkLabel(
        route_creation_frame,
        text="",
        font=FONT_SMALL
    )
    progress_label.pack(pady=(0, PADDING))

    # Gas Station Search Frame
    gas_station_frame = ctk.CTkFrame(left_column)
    gas_station_frame.pack(fill="x", pady=(0, PADDING))

    # Gas Station Search Title
    gas_station_title = ctk.CTkLabel(
        gas_station_frame,
        text="Hledat trasy podle čerpací stanice",
        font=FONT_MEDIUM
    )
    gas_station_title.pack(pady=(PADDING, 0))

    # Gas Station Entry
    gas_station_label = ctk.CTkLabel(gas_station_frame, text="Umístění čerpací stanice:", font=FONT_SMALL)
    gas_station_label.pack(anchor="w", padx=PADDING, pady=(PADDING, 0))
    
    gas_station_entry = ctk.CTkEntry(
        gas_station_frame,
        placeholder_text="Zadejte adresu čerpací stanice",
        width=300,
        font=FONT_SMALL
    )
    gas_station_entry.pack(padx=PADDING, pady=(5, PADDING))

    # Search Button
    search_station_button = ctk.CTkButton(
        gas_station_frame,
        text="Vyhledat trasy",
        command=search_routes_by_gas_station,
        width=300,
        font=FONT_SMALL
    )
    search_station_button.pack(pady=(0, PADDING))

    # Results Label
    gas_station_result_label = ctk.CTkLabel(
        gas_station_frame,
        text="",
        font=FONT_SMALL,
        justify="left",
        wraplength=280
    )
    gas_station_result_label.pack(padx=PADDING, pady=(0, PADDING))

    # Right column - Route list
    right_column = ctk.CTkFrame(content_frame, fg_color="transparent")
    right_column.pack(side="right", fill="both", expand=True)

    # Routes list frame
    routes_frame = ctk.CTkFrame(right_column)
    routes_frame.pack(fill="both", expand=True)

    routes_title = ctk.CTkLabel(
        routes_frame,
        text="Moje trasy",
        font=FONT_MEDIUM
    )
    routes_title.pack(pady=(PADDING, 0))

    # Create Treeview with modern styling
    style = ttk.Style()
    
    # Configure Treeview colors and fonts
    style.configure(
        "Minimal.Treeview",
        background="#ffffff",
        foreground="#333333",
        fieldbackground="#ffffff",
        borderwidth=0,
        font=("Roboto", 10),
        rowheight=30
    )
    
    # Configure Treeview heading style
    style.configure(
        "Minimal.Treeview.Heading",
        background="#f0f0f0",
        foreground="#333333",
        borderwidth=0,
        font=("Roboto", 10, "bold"),
        padding=10
    )
    
    # Remove the Treeview borders
    style.layout("Minimal.Treeview", [
        ('Minimal.Treeview.treearea', {'sticky': 'nswe'})
    ])
    
    # Configure selection colors
    style.map("Minimal.Treeview",
        background=[("selected", "#e6f3ff")],
        foreground=[("selected", "#000000")]
    )

    columns = ("name", "purpose", "start", "destination", "distance", "time", "fuel", "stations", "needs_fuel")
    user_routes_tree = ttk.Treeview(
        routes_frame,
        columns=columns,
        show="headings",
        style="Minimal.Treeview"
    )

    # Configure columns with shorter names and adjusted widths
    user_routes_tree.heading("name", text="Název")
    user_routes_tree.heading("purpose", text="Účel")
    user_routes_tree.heading("start", text="Odkud")
    user_routes_tree.heading("destination", text="Kam")
    user_routes_tree.heading("distance", text="Km")
    user_routes_tree.heading("time", text="Čas")
    user_routes_tree.heading("fuel", text="Palivo")
    user_routes_tree.heading("stations", text="Stanice")
    user_routes_tree.heading("needs_fuel", text="Tank")

    # Set column widths and alignment
    user_routes_tree.column("name", width=150, anchor="w")
    user_routes_tree.column("purpose", width=150, anchor="w")
    user_routes_tree.column("start", width=120, anchor="w")
    user_routes_tree.column("destination", width=120, anchor="w")
    user_routes_tree.column("distance", width=70, anchor="e")
    user_routes_tree.column("time", width=80, anchor="e")
    user_routes_tree.column("fuel", width=70, anchor="e")
    user_routes_tree.column("stations", width=70, anchor="center")
    user_routes_tree.column("needs_fuel", width=70, anchor="center")

    # Add alternating row colors
    user_routes_tree.tag_configure('oddrow', background='#f5f5f5')
    user_routes_tree.tag_configure('evenrow', background='#ffffff')

    # Add scrollbar
    scrollbar = ctk.CTkScrollbar(routes_frame, command=user_routes_tree.yview)
    scrollbar.pack(side="right", fill="y")
    user_routes_tree.configure(yscrollcommand=scrollbar.set)

    user_routes_tree.pack(fill="both", expand=True, padx=PADDING, pady=PADDING)

    # Buttons frame
    buttons_frame = ctk.CTkFrame(routes_frame, fg_color="transparent")
    buttons_frame.pack(fill="x", padx=PADDING, pady=(0, PADDING))

    edit_button = ctk.CTkButton(
        buttons_frame,
        text="Upravit trasu",
        command=edit_route_button_click,
        width=150,
        font=FONT_SMALL
    )
    edit_button.pack(side="left", padx=5)

    delete_button = ctk.CTkButton(
        buttons_frame,
        text="Smazat trasu",
        command=delete_route_button_click,
        width=150,
        fg_color="#FF5555",
        hover_color="#FF3333",
        font=FONT_SMALL
    )
    delete_button.pack(side="left", padx=5)

    # Add the new fuel management button
    add_fuel_management_button(buttons_frame)

    # Add the new car management button
    add_car_management_button(buttons_frame) # Nové tlačítko

    # Load initial routes
    load_user_routes()

def toggle_theme():
    """Toggle between light and dark theme"""
    if ctk.get_appearance_mode() == "dark":
        ctk.set_appearance_mode("light")
    else:
        ctk.set_appearance_mode("dark")

def search_routes_by_gas_station():
    """Search for routes that pass near the specified gas station."""
    global gas_station_result_label
    
    station_location = gas_station_entry.get().strip()
    if not station_location:
        messagebox.showwarning("Upozornění", "Zadejte umístění čerpací stanice.", parent=user_root)
        return

    # Get coordinates of the gas station
    if progress_label:
        progress_label.configure(text="Hledám souřadnice čerpací stanice...")
        user_root.update_idletasks()

    station_coords_str = get_coordinates(station_location)
    if not station_coords_str:
        if progress_label:
            progress_label.configure(text="")
        return

    # Convert coordinates string to float values
    try:
        station_lat, station_lon = map(float, station_coords_str.split(','))
    except ValueError:
        messagebox.showerror("Chyba", "Nepodařilo se zpracovat souřadnice čerpací stanice.", parent=user_root)
        return

    # Function to calculate distance between two points using Haversine formula
    def haversine_distance(lat1, lon1, lat2, lon2):
        from math import radians, sin, cos, sqrt, atan2
        R = 6371  # Earth's radius in kilometers

        lat1, lon1, lat2, lon2 = map(radians, [lat1, lon1, lat2, lon2])
        dlat = lat2 - lat1
        dlon = lon2 - lon1

        a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
        c = 2 * atan2(sqrt(a), sqrt(1-a))
        return R * c

    # Search through all routes
    routes_near_station = []
    max_distance_km = 5  # Maximum distance in kilometers to consider "near" the station

    for route in all_user_routes_cached:
        if 'waypoints' in route and route['waypoints']:
            # Check each waypoint in the route
            for waypoint in route['waypoints']:
                wp_lat, wp_lon = waypoint[0], waypoint[1]
                distance = haversine_distance(station_lat, station_lon, wp_lat, wp_lon)
                
                if distance <= max_distance_km:
                    routes_near_station.append((route['name'], distance))
                    break  # Found a close enough point, no need to check other waypoints

    # Update the result label
    if routes_near_station:
        result_text = "Nalezené trasy procházející kolem čerpací stanice:\n\n"
        for route_name, distance in sorted(routes_near_station, key=lambda x: x[1]):
            result_text += f"• {route_name} (vzdálenost: {distance:.1f} km)\n"
    else:
        result_text = "Žádná trasa neprochází v okolí této čerpací stanice."

    gas_station_result_label.configure(text=result_text)
    if progress_label:
        progress_label.configure(text="")

class FuelManagementWindow:
    def __init__(self, parent, routes_data):
        self.window = ctk.CTkToplevel(parent)
        self.window.title("Správa paliva")
        self.window.geometry("800x750")  # Zvětšená výška pro lepší zobrazení
        self.routes_data = routes_data
        self.refueling_entries = []
        
        self.user_cars = database.get_cars_by_user(current_user_id)
        self.selected_car_consumption = None
        self.car_var = ctk.StringVar()
        
        self.setup_ui()
        
        # Make the window modal
        self.window.transient(parent)
        self.window.grab_set()
        
    def setup_ui(self):
        # --- Main Layout Frames ---
        top_input_frame = ctk.CTkFrame(self.window)
        top_input_frame.pack(side="top", fill="x", padx=10, pady=(5,0))
        
        # Tlačítko "Generovat report" dáme úplně dolů
        generate_btn = ctk.CTkButton(
            self.window,
            text="Generovat report",
            command=self.process_monthly_data
        )
        generate_btn.pack(side="bottom", pady=10)
        
        bottom_area_frame = ctk.CTkFrame(self.window)
        bottom_area_frame.pack(side="top", fill="both", expand=True, padx=10, pady=5)
        
        # --- Populate Top Input Frame ---
        month_frame = ctk.CTkFrame(top_input_frame)
        month_frame.pack(fill="x", pady=5)
        
        month_label = ctk.CTkLabel(month_frame, text="Měsíc:")
        month_label.pack(side="left", padx=5)
        
        date_select_frame = ctk.CTkFrame(month_frame)
        date_select_frame.pack(side="left", padx=5)
        
        months = ["Leden", "Únor", "Březen", "Duben", "Květen", "Červen", 
                 "Červenec", "Srpen", "Září", "Říjen", "Listopad", "Prosinec"]
        current_month = datetime.now().month
        self.month_combo = ctk.CTkComboBox(
            date_select_frame, 
            values=months,
            width=120
        )
        self.month_combo.set(months[current_month - 1])
        self.month_combo.pack(side="left", padx=5)
        
        current_year = datetime.now().year
        self.year_var = ctk.StringVar(value=str(current_year))
        self.year_entry = ctk.CTkEntry(
            date_select_frame,
            textvariable=self.year_var,
            width=70
        )
        self.year_entry.pack(side="left", padx=5)
        
        # Tank states frame
        tank_frame = ctk.CTkFrame(top_input_frame)
        tank_frame.pack(fill="x", pady=5)
        
        initial_tank_label = ctk.CTkLabel(tank_frame, text="Počáteční stav nádrže (L):")
        initial_tank_label.pack(side="left", padx=5)
        self.initial_tank_var = ctk.StringVar(value="61.545")
        self.initial_tank_entry = ctk.CTkEntry(tank_frame, textvariable=self.initial_tank_var, width=100)
        self.initial_tank_entry.pack(side="left", padx=5)
        
        final_tank_label = ctk.CTkLabel(tank_frame, text="Cílový stav nádrže (L):")
        final_tank_label.pack(side="left", padx=5)
        self.final_tank_var = ctk.StringVar(value="30.000")
        self.final_tank_entry = ctk.CTkEntry(tank_frame, textvariable=self.final_tank_var, width=100)
        self.final_tank_entry.pack(side="left", padx=5)
        
        # Car selection and consumption frame
        car_consumption_frame = ctk.CTkFrame(top_input_frame)
        car_consumption_frame.pack(fill="x", pady=5)
        
        car_label = ctk.CTkLabel(car_consumption_frame, text="Vozidlo:")
        car_label.pack(side="left", padx=5)
        
        self.display_consumption_label = ctk.CTkLabel(car_consumption_frame, text="Spotřeba: N/A", width=250)
        
        if self.user_cars:
            car_names = [car['name'] for car in self.user_cars]
            self.car_combo = ctk.CTkComboBox(
                car_consumption_frame,
                values=car_names,
                variable=self.car_var,
                command=self._on_car_selected,
                width=200
            )
            self.car_combo.pack(side="left", padx=5)
            self.display_consumption_label.pack(side="left", padx=10)
            if car_names:
                self.car_var.set(car_names[0])
                self._on_car_selected(car_names[0])
        else:
            no_cars_label = ctk.CTkLabel(car_consumption_frame, text="Nemáte přidaná žádná vozidla.", text_color="gray")
            no_cars_label.pack(side="left", padx=5)
            self.display_consumption_label.pack(side="left", padx=10)
            self.car_combo = None
        
        # --- Populate Bottom Area Frame ---
        self.refueling_frame = ctk.CTkScrollableFrame(bottom_area_frame, height=200)
        self.refueling_frame.pack(side="top", fill="both", expand=True, pady=(0,5))
        
        refueling_label = ctk.CTkLabel(self.refueling_frame, text="Tankování během měsíce:", font=FONT_MEDIUM)
        refueling_label.pack(pady=5)
        
        add_refueling_btn = ctk.CTkButton(
            self.refueling_frame,
            text="Přidat tankování",
            command=self.add_refueling_entry
        )
        add_refueling_btn.pack(pady=5)
        
        # Results frame for displaying optimization results
        self.results_frame = ctk.CTkFrame(bottom_area_frame)
        self.results_frame.pack(side="top", fill="both", expand=True, padx=10, pady=5)
        
        # Create solution text widget
        self.solution_text = tk.Text(self.results_frame, height=10, width=50)
        self.solution_text.pack(padx=10, pady=10, fill="both", expand=True)
        
    def update_solution_display(self, solution, stats):
        """Update the display with the optimization solution and statistics."""
        if hasattr(self, 'solution_text'):
            self.solution_text.delete("1.0", tk.END)
            
            # Display the solution
            if solution:
                for route in solution:
                    self.solution_text.insert(tk.END, f"Trasa: {route.get('route_name', 'N/A')}\n")
                    self.solution_text.insert(tk.END, f"Z: {route.get('start_location', 'N/A')} Do: {route.get('destination', 'N/A')}\n")
                    self.solution_text.insert(tk.END, f"Vzdálenost: {route.get('distance', 'N/A')} km\n")
                    self.solution_text.insert(tk.END, f"Spotřeba paliva: {route.get('fuel_consumed', 'N/A')} L\n")
                    self.solution_text.insert(tk.END, f"Stav nádrže před: {route.get('fuel_before', 'N/A')} L\n")
                    self.solution_text.insert(tk.END, f"Stav nádrže po: {route.get('fuel_after', 'N/A')} L\n\n")
            
            # Display statistics
            if stats:
                self.solution_text.insert(tk.END, "\nStatistiky:\n")
                for key, value in stats.items():
                    if isinstance(value, float):
                        formatted_value = f"{value:.2f}"
                    else:
                        formatted_value = str(value)
                    self.solution_text.insert(tk.END, f"{key}: {formatted_value}\n")

    def _on_car_selected(self, selected_car_name):
        """Handle car selection change."""
        found_car = next((car for car in self.user_cars if car['name'] == selected_car_name), None)
        if found_car:
            if found_car['car_type'] == 'combustion' and found_car.get('avg_consumption') is not None:
                self.selected_car_consumption = Decimal(str(found_car['avg_consumption']))
                self.display_consumption_label.configure(text=f"Spotřeba: {self.selected_car_consumption:.2f} L/100km")
            elif found_car['car_type'] == 'electric':
                self.selected_car_consumption = Decimal('0.0')
                self.display_consumption_label.configure(text="Elektrické vozidlo (spotřeba 0.0 L/100km)")
            else:
                self.selected_car_consumption = None
                self.display_consumption_label.configure(text="Spotřeba: Není zadána", text_color="orange")
        else:
            self.selected_car_consumption = None
            self.display_consumption_label.configure(text="Spotřeba: N/A")

    def get_selected_month_date(self):
        """Get the selected month as a datetime object."""
        months = ["Leden", "Únor", "Březen", "Duben", "Květen", "Červen", 
                 "Červenec", "Srpen", "Září", "Říjen", "Listopad", "Prosinec"]
        selected_month = months.index(self.month_combo.get()) + 1
        selected_year = int(self.year_var.get())
        return datetime(selected_year, selected_month, 1)

    def process_monthly_data(self):
        """Process monthly data and generate report."""
        try:
            if not self.user_cars or not self.car_var.get():
                messagebox.showerror("Chyba", "Vyberte prosím vozidlo.", parent=self.window)
                return

            if self.selected_car_consumption is None:
                selected_car_name = self.car_var.get()
                found_car = next((car for car in self.user_cars if car['name'] == selected_car_name), None)
                if found_car and found_car['car_type'] == 'combustion':
                    messagebox.showerror("Chyba", "Vybrané spalovací vozidlo nemá zadanou průměrnou spotřebu.\nProsím, doplňte ji ve správě vozidel.", parent=self.window)
                    return
                elif not found_car:
                    messagebox.showerror("Chyba", "Nepodařilo se nalézt vybrané vozidlo.", parent=self.window)
                    return

            avg_consumption = self.selected_car_consumption

            month_date = self.get_selected_month_date()
            initial_tank = float(self.initial_tank_var.get())
            final_tank = float(self.final_tank_var.get())
            
            if not (0 <= initial_tank <= 200 and 0 <= final_tank <= 200):
                messagebox.showerror("Chyba", "Stav nádrže musí být mezi 0 a 200 litrů.", parent=self.window)
                return

            formatted_refueling_data = {}
            for entry in self.refueling_entries:
                try:
                    date_str = entry["date_var"].get()
                    amount = float(entry["amount_var"].get())
                    location = entry["location_var"].get()

                    if not date_str or amount <= 0 or not location:
                        messagebox.showerror("Chyba", "Všechna pole tankování musí být vyplněna a množství musí být větší než 0.", parent=self.window)
                        return

                    refuel_date = datetime.strptime(date_str, "%d.%m.%Y")
                    if refuel_date.strftime("%m/%Y") != month_date.strftime("%m/%Y"):
                        messagebox.showerror("Chyba", "Datum tankování musí být ve zvoleném měsíci.", parent=self.window)
                        return

                    refuel_date_obj = refuel_date.date()
                    if refuel_date_obj not in formatted_refueling_data:
                        formatted_refueling_data[refuel_date_obj] = []
                    formatted_refueling_data[refuel_date_obj].append({
                        "amount": amount,
                        "location": location
                    })
                except ValueError as e:
                    messagebox.showerror("Chyba v datech tankování", f"Neplatný formát dat u tankování: {str(e)}. Zkontrolujte datum (DD.MM.RRRR) a množství.", parent=self.window)
                    return

            print(f"\nDEBUG - Formatted refueling data:")
            print(formatted_refueling_data)

            try:
                result = integrate_with_ui(
                    initial_tank=initial_tank,
                    target_tank=final_tank,
                    start_date=month_date,
                    refueling_data=formatted_refueling_data,
                    consumption_rate=avg_consumption
                )
            except Exception as e:
                messagebox.showerror("Chyba", f"Chyba při volání optimalizace: {str(e)}", parent=self.window)
                return
            
            if result and result.get('success'):
                stats = {
                    'total_distance': result['statistics']['total_distance'],
                    'total_fuel_used': result['statistics']['total_consumption'],
                    'total_fuel_added': result['statistics']['total_refueled'],
                    'net_fuel_change': result['statistics']['total_refueled'] - result['statistics']['total_consumption'],
                    'execution_time': result['statistics'].get('execution_time', 0.0)
                }
                self.update_solution_display(result['solution'], stats)
                self.generate_monthly_report(month_date, initial_tank, final_tank, avg_consumption, formatted_refueling_data)
            else:
                error_msg = result.get('error', 'Neznámá chyba') if isinstance(result, dict) else str(result)
                messagebox.showerror(
                    "Chyba",
                    f"Nepodařilo se najít řešení: {error_msg}\n\nDetail: {result}", parent=self.window
                )
        except ValueError as e:
            messagebox.showerror("Chyba", f"Neplatný formát dat: {str(e)}", parent=self.window)
        except Exception as e:
            messagebox.showerror("Chyba", f"Nastala neočekávaná chyba: {str(e)}", parent=self.window)
            print(f"Error details: {str(e)}")

    def add_refueling_entry(self):
        """Add a new refueling entry to the form."""
        entry_frame = ctk.CTkFrame(self.refueling_frame)
        entry_frame.pack(fill="x", padx=5, pady=2)

        # Date entry with label
        date_label = ctk.CTkLabel(entry_frame, text="Datum:", font=FONT_SMALL)
        date_label.pack(side="left", padx=(5, 2))
        
        date_var = ctk.StringVar()
        date_entry = DateEntry(
            entry_frame,
            width=12,
            background='darkblue',
            foreground='white',
            borderwidth=2,
            date_pattern='dd.mm.yyyy',
            textvariable=date_var
        )
        date_entry.pack(side="left", padx=(0, 10))

        # Amount entry with label
        amount_label = ctk.CTkLabel(entry_frame, text="Množství (L):", font=FONT_SMALL)
        amount_label.pack(side="left", padx=(5, 2))
        
        amount_var = ctk.StringVar()
        amount_entry = ctk.CTkEntry(
            entry_frame,
            width=80,
            textvariable=amount_var
        )
        amount_entry.pack(side="left", padx=(0, 10))

        # Location entry with label
        location_label = ctk.CTkLabel(entry_frame, text="Místo tankování:", font=FONT_SMALL)
        location_label.pack(side="left", padx=(5, 2))
        
        location_var = ctk.StringVar()
        location_entry = ctk.CTkEntry(
            entry_frame,
            width=200,
            textvariable=location_var
        )
        location_entry.pack(side="left", padx=(0, 10))

        # Remove button
        remove_btn = ctk.CTkButton(
            entry_frame,
            text="Odebrat",
            command=lambda: self.remove_refueling_entry(entry_frame),
            width=80,
            fg_color="#FF5555",
            hover_color="#FF3333"
        )
        remove_btn.pack(side="left", padx=5)

        # Store entry data
        self.refueling_entries.append({
            "frame": entry_frame,
            "date_var": date_var,
            "amount_var": amount_var,
            "location_var": location_var
        })

    def remove_refueling_entry(self, entry_frame):
        """Remove a refueling entry from the form."""
        # Find and remove entry from the list
        self.refueling_entries = [
            entry for entry in self.refueling_entries 
            if entry["frame"] != entry_frame
        ]
        # Destroy the frame
        entry_frame.destroy()

    def generate_monthly_report(self, month_date, initial_tank, final_tank, avg_consumption, refueling_data):
        """Generate and save a monthly report in Excel format."""
        try:
            # Create a new workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Kniha jízd"

            # Define styles
            header_font = Font(bold=True)
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            pink_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Header section
            ws['A1'] = "Kniha jízd"
            ws['B1'] = "Auto:"
            ws['C1'] = self.car_var.get()
            ws['C1'].fill = yellow_fill

            ws['F1'] = "Jméno řidiče:"
            ws['G1'] = "Jiří Kvasnička"
            ws['G1'].fill = yellow_fill

            ws['K1'] = "Principal engineering s.r.o."

            ws['B2'] = "SPZ:"
            ws['C2'] = "EL 148AH"
            ws['C2'].fill = yellow_fill

            ws['F2'] = "Průměrná spotřeba podle TP:"
            ws['G2'] = f"{float(avg_consumption):.1f}"
            ws['G2'].fill = yellow_fill

            # Ensure month_date is datetime
            if isinstance(month_date, str):
                month_date = datetime.strptime(month_date, "%Y-%m-%d")
            elif isinstance(month_date, date) and not isinstance(month_date, datetime):
                month_date = datetime.combine(month_date, datetime.min.time())

            # Date range
            ws['A3'] = "Sledované období od"
            ws['D3'] = month_date.strftime("%d.%m.%Y")
            ws['A4'] = "Sledované období do"
            month_end = (month_date.replace(day=1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)
            if isinstance(month_end, date) and not isinstance(month_end, datetime):
                month_end = datetime.combine(month_end, datetime.max.time())
            ws['D4'] = month_end.strftime("%d.%m.%Y")

            # Calculate total distance from selected routes
            total_distance = sum(route['distance'] for route in self.routes_data)
            
            # Tachometer states - calculate based on total distance
            initial_tachometer = 94439  # This should come from database or user input
            final_tachometer = initial_tachometer + int(total_distance)
            
            ws['A5'] = "Počáteční stav tachometru:"
            ws['B5'] = str(initial_tachometer)
            ws['B5'].fill = yellow_fill

            ws['A6'] = "Konečný stav tachometru:"
            ws['B6'] = str(final_tachometer)
            ws['B6'].fill = yellow_fill

            # Fuel states
            ws['A7'] = "Stav nádrže k poslednímu dni předch.měsíce (l,Kč/l)"
            ws['B7'] = f"{initial_tank:.0f}"
            ws['B7'].fill = yellow_fill

            ws['A8'] = "Stav nádrže k poslednímu dni v aktuálním měsíci (l, Kč/l)"
            ws['B8'] = f"{final_tank:.0f}"
            ws['B8'].fill = yellow_fill

            # Right side information
            ws['E3'] = "Počet km - tachometr"
            ws['F3'] = str(int(total_distance))
            ws['F3'].fill = yellow_fill

            ws['E4'] = "Počet km - kniha jízd"
            ws['F4'] = str(int(total_distance))
            ws['F4'].fill = yellow_fill

            ws['E5'] = "Spotřeba PHM pro soukromé jízdy v tuz. (Kč)"
            ws['F5'] = "0,00"

            ws['E6'] = "Spotřeba PHM pro soukromé jízdy v zahr.(Kč)"
            ws['F6'] = "0,00"

            # Calculate average consumption for the period
            total_fuel_used = sum(route['fuel_consumed'] for route in self.routes_data)
            avg_period_consumption = (total_fuel_used * 100) / total_distance if total_distance > 0 else 0

            ws['E7'] = "Průměrná spotřeba v období: ( l/100 km)"
            ws['F7'] = f"{avg_period_consumption:.2f}"
            ws['F7'].fill = pink_fill

            ws['E8'] = "Průměrná cena PHM tuz.(Kč bez DPH)"
            ws['F8'] = "0,00"

            ws['E9'] = "Zahraniční cena"
            ws['F9'] = "0,00"

            # Routes table headers
            current_row = 11
            headers = ["Trasa", "Účel", "Datum", "Km", "služebně", "soukr.tuz", "soukr.zahr", 
                      "počet litrů/kW (ČR)", "Kč bez DPH (ČR)", "počet litrů (zahr)", "cena v Kč (zahr.)"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = header
                cell.font = header_font
                cell.border = border

            # Add routes data
            current_row += 1
            
            # Ensure all dates are datetime before sorting
            routes_to_sort = []
            for route in self.routes_data:
                route_copy = route.copy()
                if isinstance(route_copy['date'], str):
                    route_copy['date'] = datetime.strptime(route_copy['date'], "%Y-%m-%d")
                elif isinstance(route_copy['date'], date) and not isinstance(route_copy['date'], datetime):
                    route_copy['date'] = datetime.combine(route_copy['date'], datetime.min.time())
                routes_to_sort.append(route_copy)
                
            # Sort routes by date
            sorted_routes = sorted(routes_to_sort, key=lambda x: x['date'])
            
            for route in sorted_routes:
                ws.cell(row=current_row, column=1, value=f"{route['start_location']} - {route['destination']}")
                ws.cell(row=current_row, column=2, value=route['purpose'])
                ws.cell(row=current_row, column=3, value=route['date'].strftime("%A %d. %B %Y"))
                ws.cell(row=current_row, column=4, value=route['distance'])
                ws.cell(row=current_row, column=5, value=route['distance'])  # All km are business
                ws.cell(row=current_row, column=8, value=route['fuel_consumed'])
                current_row += 1

            # Add totals row
            ws.cell(row=current_row, column=1, value="Celkem")
            ws.cell(row=current_row, column=4, value=f"=SUM(D12:D{current_row-1})")
            ws.cell(row=current_row, column=5, value=f"=SUM(E12:E{current_row-1})")
            ws.cell(row=current_row, column=8, value=f"=SUM(H12:H{current_row-1})")

            # Adjust column widths
            ws.column_dimensions['A'].width = 35
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 15
            ws.column_dimensions['I'].width = 15
            ws.column_dimensions['J'].width = 15
            ws.column_dimensions['K'].width = 15

            # Create exports directory if it doesn't exist
            ensure_exports_dir()
            
            # Save the file
            month_year = month_date.strftime("%Y-%m")
            filename = f"kniha_jizd_{month_year}.xlsx"
            filepath = os.path.join(get_exports_path(), filename)
            
            wb.save(filepath)
            
            # Show success message and ask if user wants to open the file
            if messagebox.askyesno(
                "Report vytvořen",
                f"Report byl úspěšně vytvořen jako {filename}.\nChcete ho otevřít?",
                parent=self.window
            ):
                try:
                    os.startfile(filepath)
                except Exception as e:
                    messagebox.showwarning(
                        "Upozornění",
                        f"Report byl vytvořen, ale nepodařilo se ho automaticky otevřít.\nNajdete ho ve složce exports.\nDetail: {str(e)}",
                        parent=self.window
                    )

        except Exception as e:
            messagebox.showerror(
                "Chyba",
                f"Nepodařilo se vygenerovat report: {str(e)}",
                parent=self.window
            )
            print(f"Error generating report: {str(e)}")

class CarEditDialog:
    def __init__(self, parent, car_data=None, user_id=None):
        self.window = ctk.CTkToplevel(parent)
        self.window.title("Přidat/Upravit vozidlo")
        self.window.geometry("500x400") # Mírně snížena výška po odstranění poznámek
        self.car_data = car_data
        self.user_id = user_id
        self.result = None

        # Definice StringVar
        self.name_var = ctk.StringVar()
        self.car_type_var = ctk.StringVar(value="Spalovací") # Výchozí hodnota
        self.consumption_var = ctk.StringVar()

        self.setup_ui()
        
        # Naplnění polí, pokud editujeme (přesunuto sem, aby se provedlo po setup_ui)
        if self.car_data:
            self.name_var.set(self.car_data.get('name', ''))
            car_type_from_db = self.car_data.get('car_type', 'combustion')
            self.car_type_var.set("Elektrické" if car_type_from_db == 'electric' else "Spalovací")
            
            if car_type_from_db == 'combustion' and self.car_data.get('avg_consumption') is not None:
                self.consumption_var.set(str(self.car_data.get('avg_consumption', '')))
            else:
                self.consumption_var.set('') # Vyčistí, pokud je elektrické nebo nemá spotřebu
            
            # Aktualizujeme zobrazení pole spotřeby na základě načtených dat
            self.toggle_consumption_field() 

        self.window.transient(parent)
        self.window.grab_set()
        
    def setup_ui(self):
        form_frame = ctk.CTkFrame(self.window, fg_color="transparent")
        form_frame.pack(padx=PADDING, pady=PADDING, fill="both", expand=True)

        # Name entry
        name_label = ctk.CTkLabel(form_frame, text="Název vozidla (např. Škoda Octavia):", font=FONT_NORMAL)
        name_label.pack(pady=(0, 5), anchor="w")
        self.name_entry = ctk.CTkEntry(form_frame, textvariable=self.name_var, width=350) # Použití name_var
        self.name_entry.pack(pady=(0, 10), fill="x")

        # Car type (combobox)
        type_label = ctk.CTkLabel(form_frame, text="Typ vozidla:", font=FONT_NORMAL)
        type_label.pack(pady=(0, 5), anchor="w")
        self.car_type_var = ctk.StringVar(value="Spalovací")
        self.car_type_combo = ctk.CTkComboBox(
            form_frame, 
            values=["Spalovací", "Elektrické"],
            variable=self.car_type_var,
            command=self.toggle_consumption_field,
            width=350
        )
        self.car_type_combo.pack(pady=(0, 10), fill="x")

        # Average consumption (conditionally visible)
        self.consumption_label = ctk.CTkLabel(form_frame, text="Průměrná spotřeba (L/100km):", font=FONT_NORMAL)
        self.consumption_entry = ctk.CTkEntry(form_frame, textvariable=self.consumption_var, width=350) # Použití consumption_var
        # Zobrazíme/skryjeme podle počáteční hodnoty car_type_var

        # Notes entry
        # notes_label = ctk.CTkLabel(form_frame, text="Poznámky:", font=FONT_NORMAL)
        # notes_label.pack(pady=(0, 5), anchor="w")
        # self.notes_entry = ctk.CTkTextbox(form_frame, height=80)
        # self.notes_entry.pack(pady=(0, 15), fill="x")
        
        # Populate fields if editing - PŘESUNUTO do __init__ po setup_ui()
        # if self.car_data:
        # self.name_entry.insert(0, self.car_data.get('name', ''))
            # self.notes_entry.insert("1.0", self.car_data.get('notes', ''))
        
        self.toggle_consumption_field() # Ensure correct initial state

        # Buttons frame
        button_frame = ctk.CTkFrame(self.window, fg_color="transparent")
        button_frame.pack(pady=PADDING, side="bottom")
        
        save_button = ctk.CTkButton(button_frame, text="Uložit", command=self.save, width=120, font=FONT_SMALL)
        save_button.pack(side="left", padx=10)
        cancel_button = ctk.CTkButton(button_frame, text="Zrušit", command=self.cancel, width=120, fg_color="gray", hover_color="darkgray", font=FONT_SMALL)
        cancel_button.pack(side="left", padx=10)

    def toggle_consumption_field(self, event=None): # event je pro command comboboxu
        if self.car_type_var.get() == "Spalovací":
            self.consumption_label.pack(pady=(10, 5), anchor="w") # Znovu packujeme, aby se zobrazilo
            self.consumption_entry.pack(pady=(0, 10), fill="x")
        else:
            self.consumption_label.pack_forget()
            self.consumption_entry.pack_forget()
            self.consumption_entry.delete(0, 'end') # Vymažeme hodnotu, pokud se přepne na elektrické
        
    def save(self):
        name = self.name_var.get().strip() # Použití name_var
        car_type_ui = self.car_type_var.get()
        car_type_db = 'electric' if car_type_ui == "Elektrické" else 'combustion'
        avg_consumption_str = self.consumption_var.get().strip() # Použití consumption_var
        # notes = self.notes_entry.get("1.0", "end-1c").strip() # Poznámky jsou odstraněny

        if not name:
            messagebox.showerror("Chyba", "Název vozidla je povinný.", parent=self.window)
            return

        avg_consumption = None
        if car_type_db == 'combustion':
            if not avg_consumption_str:
                messagebox.showerror("Chyba", "Průměrná spotřeba je pro spalovací vozidlo povinná.", parent=self.window)
                return
            try:
                avg_consumption = Decimal(avg_consumption_str.replace(',', '.'))
                if avg_consumption <= 0:
                    raise ValueError("Spotřeba musí být kladné číslo.")
            except (ValueError, TypeError) as e:
                messagebox.showerror("Chyba", f"Neplatná hodnota pro průměrnou spotřebu: {e}", parent=self.window)
                return
        
        if self.car_data: # Editace existujícího
            success, message = database.update_car(
                self.car_data['id'], self.car_data['user_id'], name,
                car_type_db, avg_consumption # Bez notes
            )
        else: # Přidání nového
            if not self.user_id:
                messagebox.showerror("Chyba", "Interní chyba: Chybí ID uživatele.", parent=self.window)
                return            
            success, message_or_id = database.add_car(
                self.user_id, name, car_type_db, avg_consumption # Bez notes
            )
            # 'message_or_id' je buď chybová zpráva (string) nebo ID nového auta (int)
            if success:
                message = "Vozidlo úspěšně přidáno."
            else:
                message = message_or_id

        if success:
            messagebox.showinfo("Úspěch", message, parent=self.window)
            self.result = True # Signalizujeme, že se má obnovit seznam
            self.window.destroy()
        else:
            messagebox.showerror("Chyba", message, parent=self.window)
            
    def cancel(self):
        self.window.destroy()

class CarManagementWindow:
    def __init__(self, parent):
        self.parent = parent # Uložíme si referenci na rodičovské okno (user_root)
        self.window = ctk.CTkToplevel(parent)
        self.window.title("Správa vozidel")
        self.window.geometry("900x600")
        self.window.transient(parent)
        self.window.grab_set()

        self.cars_tree = None
        self.setup_ui()
        self.load_cars()

    def setup_ui(self):
        main_frame = ctk.CTkFrame(self.window)
        main_frame.pack(fill="both", expand=True, padx=PADDING, pady=PADDING)

        title_label = ctk.CTkLabel(main_frame, text="Moje vozidla", font=FONT_LARGE)
        title_label.pack(pady=(0, PADDING))

        # Treeview for cars
        tree_frame = ctk.CTkFrame(main_frame)
        tree_frame.pack(fill="both", expand=True, pady=(0, PADDING))

        columns = ("name", "type", "consumption")
        self.cars_tree = ttk.Treeview(
            tree_frame,
            columns=columns,
            show="headings",
            style="Minimal.Treeview" # Použijeme existující styl
        )
        self.cars_tree.heading("name", text="Název vozidla")
        self.cars_tree.heading("type", text="Typ")
        self.cars_tree.heading("consumption", text="Spotřeba (L/100km)")

        self.cars_tree.column("name", width=250, anchor="w") # Širší název
        self.cars_tree.column("type", width=150, anchor="center")
        self.cars_tree.column("consumption", width=200, anchor="e")

        scrollbar = ctk.CTkScrollbar(tree_frame, command=self.cars_tree.yview)
        scrollbar.pack(side="right", fill="y")
        self.cars_tree.configure(yscrollcommand=scrollbar.set)
        self.cars_tree.pack(fill="both", expand=True)

        # Buttons frame
        buttons_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        buttons_frame.pack(fill="x", pady=(PADDING, 0))

        add_button = ctk.CTkButton(buttons_frame, text="Přidat vozidlo", command=self.add_car_dialog, font=FONT_SMALL)
        add_button.pack(side="left", padx=5)

        edit_button = ctk.CTkButton(buttons_frame, text="Upravit vozidlo", command=self.edit_car_dialog, font=FONT_SMALL)
        edit_button.pack(side="left", padx=5)

        delete_button = ctk.CTkButton(buttons_frame, text="Smazat vozidlo", command=self.delete_selected_car, fg_color="#FF5555", hover_color="#FF3333", font=FONT_SMALL)
        delete_button.pack(side="left", padx=5)

    def load_cars(self):
        for item in self.cars_tree.get_children():
            self.cars_tree.delete(item)
        
        cars = database.get_cars_by_user(current_user_id) # Předpokládá globální current_user_id
        if cars:
            for car in cars:
                consumption_display = f"{car.get('avg_consumption'):.2f}" if car.get('avg_consumption') else "N/A"
                car_type_display = "Elektrické" if car.get('car_type') == 'electric' else "Spalovací"
                values = (
                    car.get('name', ''),
                    car_type_display,
                    consumption_display,
                )
                self.cars_tree.insert("", "end", values=values, tags=(car['id'],))

    def get_selected_car_id(self):
        selected_item = self.cars_tree.focus()
        if not selected_item:
            messagebox.showwarning("Žádný výběr", "Nejprve vyberte vozidlo ze seznamu.", parent=self.window)
            return None
        return self.cars_tree.item(selected_item, "tags")[0]

    def add_car_dialog(self):
        dialog = CarEditDialog(self.window, user_id=current_user_id) # Předáme user_id pro nové auto
        self.window.wait_window(dialog.window)
        if dialog.result: # Pokud bylo úspěšně uloženo
            self.load_cars()

    def edit_car_dialog(self):
        car_id = self.get_selected_car_id()
        if not car_id:
            return
        
        # Načteme kompletní data o autě z DB, abychom měli vše pro dialog
        car_data = database.get_car_by_id(car_id, current_user_id)
        if not car_data:
            messagebox.showerror("Chyba", "Nepodařilo se načíst data vozidla.", parent=self.window)
            return

        dialog = CarEditDialog(self.window, car_data=car_data)
        self.window.wait_window(dialog.window)
        if dialog.result:
            self.load_cars()

    def delete_selected_car(self):
        car_id = self.get_selected_car_id()
        if not car_id:
            return

        if messagebox.askyesno("Smazat vozidlo", "Opravdu chcete smazat vybrané vozidlo?", parent=self.window):
            success, message = database.delete_car(car_id, current_user_id)
            if success:
                messagebox.showinfo("Úspěch", message, parent=self.window)
                self.load_cars()
            else:
                messagebox.showerror("Chyba", message, parent=self.window)

def add_car_management_button(parent_frame):
    """Přidá tlačítko pro otevření okna správy vozidel."""
    def open_car_management():
        CarManagementWindow(parent_frame.winfo_toplevel()) # Jako parent předáme hlavní okno aplikace

    car_management_btn = ctk.CTkButton(
        parent_frame,
        text="Správa vozidel",
        command=open_car_management,
        width=140, # Trochu širší
        font=FONT_SMALL
    )
    car_management_btn.pack(side="left", padx=5) # Changed from right to left

def add_fuel_management_button(parent_frame):
    """Přidá tlačítko pro otevření okna správy tankování."""
    def open_fuel_management():
        # Získáme všechny trasy pro aktuální měsíc
        current_month = datetime.now().month
        current_year = datetime.now().year
        month_routes = [route for route in all_user_routes_cached 
                       if isinstance(route.get('date'), (date, datetime)) and 
                       route['date'].month == current_month and 
                       route['date'].year == current_year]
        
        # Otevřeme okno správy paliva s aktuálními trasami
        FuelManagementWindow(parent_frame.winfo_toplevel(), month_routes)

    fuel_management_btn = ctk.CTkButton(
        parent_frame,
        text="Správa tankování",
        command=open_fuel_management,
        width=140,
        font=FONT_SMALL
    )
    fuel_management_btn.pack(side="left", padx=5) # Changed from right to left